from __future__ import annotations

import os
import json
import random
import re
import io
import zipfile
from contextlib import contextmanager
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Generator, Optional
from urllib.parse import quote
from urllib.request import Request as UrlRequest, urlopen
from urllib.error import HTTPError, URLError

from fastapi import Depends, FastAPI, HTTPException, Query, Request, status
from fastapi.responses import FileResponse, Response, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from pydantic_settings import BaseSettings, SettingsConfigDict
from sqlalchemy import Boolean, CheckConstraint, Date, DateTime, ForeignKey, Integer, String, UniqueConstraint, create_engine, delete, func, select, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, relationship, sessionmaker
from starlette.middleware.sessions import SessionMiddleware

from openpyxl import Workbook

try:
    from docx import Document
except ImportError:  # pragma: no cover
    Document = None  # type: ignore[misc, assignment]

try:
    from passlib.context import CryptContext
except Exception:  # pragma: no cover
    CryptContext = None  # type: ignore


class Settings(BaseSettings):
    model_config = SettingsConfigDict(env_file=None, extra="ignore")

    DATABASE_PATH: str = "data/app.sqlite"
    EXPORTS_DIR: str = "exports"
    SESSION_SECRET: str = ""

    BOOTSTRAP_ADMIN_USERNAME: str = ""
    BOOTSTRAP_ADMIN_PASSWORD: str = ""
    BOOTSTRAP_ADMIN_FULLNAME: str = ""

    PORT: int = 8000
    CORS_ALLOW_ORIGINS: str = ""  # comma-separated
    N8N_WEBHOOK_URL: str = ""
    N8N_WEBHOOK_TIMEOUT_SEC: int = 5


settings = Settings()


def _pwd_context() -> "CryptContext":
    if CryptContext is None:
        raise RuntimeError("passlib is required for password hashing")
    return CryptContext(schemes=["bcrypt"], deprecated="auto")


pwd_context = _pwd_context()


def hash_password(password: str) -> str:
    return pwd_context.hash(password)


def verify_password(password: str, password_hash: str) -> bool:
    return pwd_context.verify(password, password_hash)


class Base(DeclarativeBase):
    pass


class User(Base):
    __tablename__ = "users"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    username: Mapped[str] = mapped_column(String(64), unique=True, nullable=False, index=True)
    full_name: Mapped[str] = mapped_column(String(200), nullable=False)
    role: Mapped[str] = mapped_column(String(16), nullable=False, index=True)  # "admin" | "support"
    password_hash: Mapped[str] = mapped_column(String(200), nullable=False)
    is_active_for_duties: Mapped[bool] = mapped_column(Boolean, nullable=False, default=True)

    duty_assignments: Mapped[list["DutyAssignment"]] = relationship(
        back_populates="support_user", cascade="all,delete"
    )
    reports: Mapped[list["DailyReport"]] = relationship(back_populates="support_user", cascade="all,delete")


class DutyAssignment(Base):
    __tablename__ = "duty_assignments"

    date: Mapped[date] = mapped_column(Date, primary_key=True)
    slot: Mapped[int] = mapped_column(Integer, primary_key=True)
    support_user_id: Mapped[int] = mapped_column(ForeignKey("users.id"), nullable=False)

    __table_args__ = (
        CheckConstraint("slot >= 0 AND slot <= 10", name="ck_duty_slot_range"),
    )

    support_user: Mapped[User] = relationship(back_populates="duty_assignments")


class DutySwapRequest(Base):
    __tablename__ = "duty_swap_requests"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    date: Mapped[date] = mapped_column(Date, nullable=False, index=True)
    from_slot: Mapped[int] = mapped_column(Integer, nullable=False)
    to_slot: Mapped[int] = mapped_column(Integer, nullable=False)
    requester_user_id: Mapped[int] = mapped_column(ForeignKey("users.id"), nullable=False, index=True)
    target_user_id: Mapped[int] = mapped_column(ForeignKey("users.id"), nullable=False, index=True)
    message: Mapped[str] = mapped_column(String(500), nullable=False)
    status: Mapped[str] = mapped_column(String(16), nullable=False, default="pending", index=True)
    created_at: Mapped[datetime] = mapped_column(DateTime, nullable=False, default=datetime.utcnow, index=True)

    __table_args__ = (
        CheckConstraint("from_slot >= 0 AND from_slot <= 10", name="ck_swap_from_slot_range"),
        CheckConstraint("to_slot >= 0 AND to_slot <= 10", name="ck_swap_to_slot_range"),
        CheckConstraint("status IN ('pending','accepted','rejected')", name="ck_swap_status"),
    )


class DailyReport(Base):
    __tablename__ = "daily_reports"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    date: Mapped[date] = mapped_column(Date, nullable=False, index=True)
    support_user_id: Mapped[int] = mapped_column(ForeignKey("users.id"), nullable=False, index=True)
    status: Mapped[str] = mapped_column(String(16), nullable=False, default="draft", index=True)
    finalized_at: Mapped[Optional[datetime]] = mapped_column(DateTime, nullable=True)

    support_user: Mapped[User] = relationship(back_populates="reports")
    entries: Mapped[list["ReportEntry"]] = relationship(
        back_populates="report", cascade="all,delete-orphan", order_by="ReportEntry.id"
    )

    __table_args__ = (  # type: ignore[assignment]
        # Only one report per support user per date.
        UniqueConstraint("date", "support_user_id", name="uq_daily_report_date_employee"),
        CheckConstraint("status IN ('draft','final')", name="ck_daily_report_status"),
    )


class ReportEntry(Base):
    __tablename__ = "report_entries"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    report_id: Mapped[int] = mapped_column(ForeignKey("daily_reports.id"), nullable=False, index=True)
    minutes: Mapped[int] = mapped_column(Integer, nullable=False)
    description: Mapped[str] = mapped_column(String(2000), nullable=False)

    report: Mapped[DailyReport] = relationship(back_populates="entries")


engine = None
SessionLocal = None  # type: ignore


def get_engine():
    global engine, SessionLocal
    if engine is not None:
        return engine

    db_path = Path(settings.DATABASE_PATH).expanduser()
    if not db_path.is_absolute():
        db_path = (Path.cwd() / db_path).resolve()
    db_path.parent.mkdir(parents=True, exist_ok=True)

    # sqlite URL; sqlite will create the file automatically
    url = f"sqlite:///{str(db_path)}"
    engine = create_engine(
        url,
        connect_args={"check_same_thread": False},
    )
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    return engine


def init_db() -> None:
    eng = get_engine()
    Base.metadata.create_all(bind=eng)


@contextmanager
def db_session() -> Generator:
    if SessionLocal is None:
        get_engine()
    db = SessionLocal()  # type: ignore[misc]
    try:
        yield db
    finally:
        db.close()


def get_db():
    with db_session() as db:
        yield db


SLOT_START_HOUR = 7
SLOT_COUNT = 11
SLOT_MAX_INDEX = SLOT_COUNT - 1


def slot_start_time_str(slot: int) -> str:
    return (time(hour=SLOT_START_HOUR + slot, minute=0)).strftime("%H:%M")


def slot_time_delta(slot: int) -> timedelta:
    return timedelta(hours=slot)


def duty_slot_for_dt(dt: datetime) -> tuple[date, int] | tuple[None, None]:
    slot = dt.hour - SLOT_START_HOUR
    if slot < 0 or slot > SLOT_MAX_INDEX:
        return None, None
    return dt.date(), slot


def ensure_support_user(db, user_id: int) -> User:
    u = db.get(User, user_id)
    if not u or u.role != "support":
        raise HTTPException(status_code=400, detail="User is not a support user")
    return u


def get_current_user(request: Request, db=Depends(get_db)) -> User:
    user_id = request.session.get("user_id")
    if not user_id:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Not authenticated")
    user = db.get(User, int(user_id))
    if not user:
        request.session.clear()
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Not authenticated")
    return user


def require_admin(current_user: User = Depends(get_current_user)) -> User:
    if current_user.role != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin only")
    return current_user


def require_support_or_admin(current_user: User = Depends(get_current_user)) -> User:
    if current_user.role not in ("support", "admin"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")
    return current_user


class LoginRequest(BaseModel):
    username: str = Field(min_length=1, max_length=64)
    password: str = Field(min_length=1, max_length=200)


class CreateSupportUserRequest(BaseModel):
    username: str = Field(min_length=1, max_length=64)
    full_name: str = Field(min_length=1, max_length=200)
    password: str = Field(min_length=1, max_length=200)


class AdminChangePasswordRequest(BaseModel):
    new_password: str = Field(min_length=8, max_length=200)


class AdminDutyStatusRequest(BaseModel):
    is_active_for_duties: bool


class AdminUpdateUserRequest(BaseModel):
    username: str = Field(min_length=1, max_length=64)
    full_name: str = Field(min_length=1, max_length=200)


class SelfChangePasswordRequest(BaseModel):
    old_password: str = Field(min_length=1, max_length=200)
    new_password: str = Field(min_length=8, max_length=200)


class SelfUpdateProfileRequest(BaseModel):
    full_name: str = Field(min_length=1, max_length=200)


class EmployeeExitInstructionRequest(BaseModel):
    fio: str = Field(min_length=1, max_length=300)
    login: str = Field(min_length=1, max_length=128)
    password: str = Field(min_length=1, max_length=200)
    domain: str = Field(min_length=1, max_length=128)


class EmployeeExitInstructionOut(BaseModel):
    text: str


class UserOut(BaseModel):
    id: int
    username: str
    full_name: str
    role: str
    is_active_for_duties: bool = True


class UserMeOut(UserOut):
    pass


class LogoutOut(BaseModel):
    ok: bool = True


class DutySlotOut(BaseModel):
    slot: int
    start_time: str
    user: Optional[UserOut] = None


class DutiesOut(BaseModel):
    date: date
    slots: list[DutySlotOut]


class DutiesGenerateRequest(BaseModel):
    start_date: date
    end_date: date
    overwrite: bool = False


class DutiesBatchSlot(BaseModel):
    slot: int = Field(ge=0, le=SLOT_MAX_INDEX)
    user_id: int = Field(ge=1)


class DutiesBatchRequest(BaseModel):
    date: date
    assignments: list[DutiesBatchSlot]


class DutySwapCreateRequest(BaseModel):
    date: date
    from_slot: int = Field(ge=0, le=SLOT_MAX_INDEX)
    to_slot: int = Field(ge=0, le=SLOT_MAX_INDEX)


class DutySwapOut(BaseModel):
    id: int
    date: date
    from_slot: int
    to_slot: int
    requester_user_id: int
    target_user_id: int
    message: str
    status: str
    created_at: datetime


class DutySwapDecisionRequest(BaseModel):
    action: str = Field(pattern="^(accept|reject)$")


class ReportEntryIn(BaseModel):
    minutes: int = Field(ge=0, le=24 * 60)
    description: str = Field(min_length=1, max_length=2000)


class CreateOrGetReportRequest(BaseModel):
    date: date
    employee_id: Optional[int] = Field(default=None, ge=1)


class UpdateReportRequest(BaseModel):
    date: Optional[date] = None
    employee_id: Optional[int] = None
    entries: list[ReportEntryIn]


class ReportEntryOut(BaseModel):
    minutes: int
    description: str


class DailyReportOut(BaseModel):
    report_id: int
    date: date
    employee_id: int
    employee: UserOut
    status: str
    finalized_at: Optional[datetime] = None
    entries: list[ReportEntryOut] = Field(default_factory=list)


class DutiesGenerateOut(BaseModel):
    start_date: date
    end_date: date
    overwrite: bool
    created_assignments: int


class ReportFinalizeOut(BaseModel):
    report_id: int
    status: str
    excel_url: str


class DutyNotificationDispatchOut(BaseModel):
    sent: bool
    reason: Optional[str] = None
    date: date
    slot: int
    start_time: str
    employee_id: Optional[int] = None
    employee_name: Optional[str] = None


EXPORT_FILENAME_RE = re.compile(r"^report_(\d+)_\d{4}-\d{2}-\d{2}\.xlsx$")


def exports_dir() -> Path:
    p = Path(settings.EXPORTS_DIR).expanduser()
    if not p.is_absolute():
        p = (Path.cwd() / p).resolve()
    p.mkdir(parents=True, exist_ok=True)
    return p


def build_employee_exit_instruction(fio: str, login: str, password: str, domain: str) -> str:
    fio = (fio or "").strip()
    login = (login or "").strip()
    password = (password or "").strip()
    domain = (domain or "").strip()
    domain_login_example = f"{domain}\\{login}"
    return (
        f"Добрый день, {fio}, я системный администратор в компании Sokolov, вам выдано оборудование.\n\n"
        "При включении ноутбука открывается bitlocker - стандартный пароль от него Sokolov2026 \n"
        f"Ваш логин - {login}, ваш пароль, при первом входе попросит сменить - {password}.\n\n"
        f"Ваш домен — {domain}.\n"
        f"Пример учётной записи в формате домена: {domain_login_example}\n\n"
        "Вход в сервисы осуществляется по доменной учётной записи.\n\n"
        "После входа в учетную запись вы можете войти в информационные ресурсы компании, почту, битрикс24. \n"
        "При входе в Битрикс у вас запросит адрес сайта - 'portal.hpdd.ru', логин и пароль от вашей доменной учётной записи.\n"
        "Для входа в Outlook также используется доменная учётная запись.\n"
        f"При входе в ZOOM нужно выбрать вход через Active Directory и ввести учётную запись в формате {domain_login_example} и пароль.\n\n"
        "Важно:\n"
        "• Папка 'Загрузки' автоматически очищается при перезагрузке.\n"
        "• Папка 'Документы' синхронизируется с сервером, для удобства при смене оборудования - данные будут синхронизированы.\n\n"
        "По всем вопросам вы можете набрать по номеру 8 800 1000 750 (добавочный уточняйте у руководителя или в службе поддержки)."
    )


def build_employee_exit_instruction_docx_bytes(fio: str, login: str, password: str, domain: str) -> bytes:
    if Document is None:
        raise HTTPException(
            status_code=503,
            detail="Генерация Word недоступна: в образе не установлен пакет python-docx. Пересоберите backend.",
        )
    text = build_employee_exit_instruction(fio, login, password, domain)
    doc = Document()
    for line in text.splitlines():
        doc.add_paragraph(line)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def attachment_filename_docx(base_name: str) -> str:
    safe = safe_export_name_part(base_name) or "employee"
    return f"instrukciya_{safe}.docx"


def report_excel_filename(report: DailyReport) -> str:
    return f"report_{report.id}_{report.date.isoformat()}.xlsx"


def surname_from_full_name(full_name: str) -> str:
    parts = [p for p in (full_name or "").strip().split() if p]
    if not parts:
        return "Сотрудник"
    return parts[0]


def safe_export_name_part(value: str) -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|]+", "_", value).strip()
    return cleaned or "Сотрудник"


def build_report_excel(report: DailyReport, user: User) -> Path:
    out_dir = exports_dir()
    filename = report_excel_filename(report)
    out_path = out_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"

    # Header
    ws["A1"] = "Support Daily Report"
    ws["A3"] = "Employee"
    ws["B3"] = f"{user.full_name} ({user.username})"
    ws["A4"] = "Date"
    ws["B4"] = report.date.isoformat()
    ws["A5"] = "Status"
    ws["B5"] = report.status
    ws["A6"] = "GeneratedAt"
    ws["B6"] = datetime.utcnow().isoformat() + "Z"

    # Entries table
    ws["A8"] = "№"
    ws["B8"] = "Minutes"
    ws["C8"] = "Description"

    total_minutes = 0
    row = 9
    for idx, entry in enumerate(report.entries, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = entry.minutes
        ws[f"C{row}"] = entry.description
        total_minutes += entry.minutes
        row += 1

    ws[f"A{row + 1}"] = "TotalMinutes"
    ws[f"B{row + 1}"] = total_minutes

    # Basic column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 80

    wb.save(out_path)
    return out_path


def report_to_out(db, report: DailyReport) -> DailyReportOut:
    employee = db.get(User, report.support_user_id)
    if not employee:
        raise HTTPException(status_code=500, detail="Employee not found")
    return DailyReportOut(
        report_id=report.id,
        date=report.date,
        employee_id=report.support_user_id,
        employee=UserOut(
            id=employee.id,
            username=employee.username,
            full_name=employee.full_name,
            role=employee.role,
        ),
        status=report.status,
        finalized_at=report.finalized_at,
        entries=[ReportEntryOut(minutes=e.minutes, description=e.description) for e in report.entries],
    )


app = FastAPI(title="Support Duty & Reports API", redirect_slashes=False)

cors_origins = [o.strip() for o in settings.CORS_ALLOW_ORIGINS.split(",") if o.strip()]
if cors_origins:
    app.add_middleware(
        CORSMiddleware,
        allow_origins=cors_origins,
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )


def bootstrap_admin_if_needed(db) -> None:
    existing_users_count = db.execute(select(func.count()).select_from(User)).scalar_one()
    if existing_users_count and existing_users_count > 0:
        return

    if not settings.BOOTSTRAP_ADMIN_USERNAME or not settings.BOOTSTRAP_ADMIN_PASSWORD or not settings.BOOTSTRAP_ADMIN_FULLNAME:
        raise RuntimeError("Bootstrap admin env vars are required: BOOTSTRAP_ADMIN_USERNAME/PASSWORD/FULLNAME")

    admin = User(
        username=settings.BOOTSTRAP_ADMIN_USERNAME,
        full_name=settings.BOOTSTRAP_ADMIN_FULLNAME,
        role="admin",
        password_hash=hash_password(settings.BOOTSTRAP_ADMIN_PASSWORD),
    )
    db.add(admin)
    db.commit()


def ensure_user_is_active_column() -> None:
    eng = get_engine()
    with eng.begin() as conn:
        cols = conn.exec_driver_sql("PRAGMA table_info(users)").fetchall()
        names = {str(c[1]) for c in cols}
        if "is_active_for_duties" not in names:
            conn.exec_driver_sql(
                "ALTER TABLE users ADD COLUMN is_active_for_duties INTEGER NOT NULL DEFAULT 1"
            )


def ensure_duty_swap_status_column() -> None:
    eng = get_engine()
    with eng.begin() as conn:
        cols = conn.exec_driver_sql("PRAGMA table_info(duty_swap_requests)").fetchall()
        names = {str(c[1]) for c in cols}
        if "status" not in names:
            conn.exec_driver_sql(
                "ALTER TABLE duty_swap_requests ADD COLUMN status TEXT NOT NULL DEFAULT 'pending'"
            )


def ensure_slot_range_constraints() -> None:
    eng = get_engine()
    with eng.begin() as conn:
        duty_sql = conn.exec_driver_sql(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='duty_assignments'"
        ).scalar_one_or_none()
        if duty_sql and "slot <= 8" in str(duty_sql):
            conn.exec_driver_sql(
                """
                CREATE TABLE duty_assignments_new (
                    date DATE NOT NULL,
                    slot INTEGER NOT NULL,
                    support_user_id INTEGER NOT NULL,
                    PRIMARY KEY (date, slot),
                    CONSTRAINT ck_duty_slot_range CHECK (slot >= 0 AND slot <= 10),
                    FOREIGN KEY(support_user_id) REFERENCES users (id)
                )
                """
            )
            conn.exec_driver_sql(
                "INSERT INTO duty_assignments_new(date, slot, support_user_id) "
                "SELECT date, slot, support_user_id FROM duty_assignments"
            )
            conn.exec_driver_sql("DROP TABLE duty_assignments")
            conn.exec_driver_sql("ALTER TABLE duty_assignments_new RENAME TO duty_assignments")

        swap_sql = conn.exec_driver_sql(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='duty_swap_requests'"
        ).scalar_one_or_none()
        if swap_sql and ("from_slot <= 8" in str(swap_sql) or "to_slot <= 8" in str(swap_sql)):
            conn.exec_driver_sql(
                """
                CREATE TABLE duty_swap_requests_new (
                    id INTEGER NOT NULL,
                    date DATE NOT NULL,
                    from_slot INTEGER NOT NULL,
                    to_slot INTEGER NOT NULL,
                    requester_user_id INTEGER NOT NULL,
                    target_user_id INTEGER NOT NULL,
                    message VARCHAR(500) NOT NULL,
                    status VARCHAR(16) NOT NULL DEFAULT 'pending',
                    created_at DATETIME NOT NULL,
                    PRIMARY KEY (id),
                    CONSTRAINT ck_swap_from_slot_range CHECK (from_slot >= 0 AND from_slot <= 10),
                    CONSTRAINT ck_swap_to_slot_range CHECK (to_slot >= 0 AND to_slot <= 10),
                    CONSTRAINT ck_swap_status CHECK (status IN ('pending','accepted','rejected')),
                    FOREIGN KEY(requester_user_id) REFERENCES users (id),
                    FOREIGN KEY(target_user_id) REFERENCES users (id)
                )
                """
            )
            conn.exec_driver_sql(
                "INSERT INTO duty_swap_requests_new(id, date, from_slot, to_slot, requester_user_id, target_user_id, message, status, created_at) "
                "SELECT id, date, from_slot, to_slot, requester_user_id, target_user_id, message, status, created_at FROM duty_swap_requests"
            )
            conn.exec_driver_sql("DROP TABLE duty_swap_requests")
            conn.exec_driver_sql("ALTER TABLE duty_swap_requests_new RENAME TO duty_swap_requests")


@app.on_event("startup")
def on_startup() -> None:
    # init DB tables + bootstrap admin user
    init_db()
    ensure_user_is_active_column()
    ensure_duty_swap_status_column()
    ensure_slot_range_constraints()
    if not settings.SESSION_SECRET:
        # SessionMiddleware needs it; use explicit error to avoid confusing auth failures.
        raise RuntimeError("SESSION_SECRET env var is required")

    with db_session() as db:
        bootstrap_admin_if_needed(db)


app.add_middleware(SessionMiddleware, secret_key=settings.SESSION_SECRET, session_cookie="session")


@app.get("/api/health")
def health() -> dict:
    return {"ok": True}


@app.post("/api/login", response_model=UserMeOut)
def login(payload: LoginRequest, request: Request, db=Depends(get_db)) -> UserMeOut:
    user = db.execute(select(User).where(User.username == payload.username)).scalar_one_or_none()
    if not user or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid credentials")

    request.session["user_id"] = str(user.id)
    return UserMeOut(id=user.id, username=user.username, full_name=user.full_name, role=user.role)


@app.post("/api/logout", response_model=LogoutOut)
def logout(request: Request) -> LogoutOut:
    request.session.clear()
    return LogoutOut(ok=True)


@app.get("/api/me", response_model=UserMeOut)
def me(current_user: User = Depends(get_current_user)) -> UserMeOut:
    return UserMeOut(id=current_user.id, username=current_user.username, full_name=current_user.full_name, role=current_user.role)


@app.post("/api/me/password")
def change_own_password(
    payload: SelfChangePasswordRequest,
    current_user: User = Depends(get_current_user),
    db=Depends(get_db),
) -> dict:
    if not verify_password(payload.old_password, current_user.password_hash):
        raise HTTPException(status_code=400, detail="Invalid old password")
    if verify_password(payload.new_password, current_user.password_hash):
        raise HTTPException(status_code=400, detail="New password must differ from old password")

    current_user.password_hash = hash_password(payload.new_password)
    db.add(current_user)
    db.commit()
    return {"ok": True}


@app.patch("/api/me/profile", response_model=UserMeOut)
def update_own_profile(
    payload: SelfUpdateProfileRequest,
    current_user: User = Depends(get_current_user),
    db=Depends(get_db),
) -> UserMeOut:
    current_user.full_name = payload.full_name.strip()
    db.add(current_user)
    db.commit()
    db.refresh(current_user)
    return UserMeOut(
        id=current_user.id,
        username=current_user.username,
        full_name=current_user.full_name,
        role=current_user.role,
    )


@app.get("/api/employee_exit/ping")
@app.get("/api/ee_instruction/ping")
def employee_exit_ping() -> dict:
    """Публичная проверка маршрута (без сессии)."""
    return {"ok": True, "employee_exit": True}


@app.post("/api/ee_instruction", response_model=EmployeeExitInstructionOut)
@app.post("/api/employee_exit/instruction", response_model=EmployeeExitInstructionOut)
@app.post("/api/employee-exit/instruction", response_model=EmployeeExitInstructionOut)
def employee_exit_instruction(
    payload: EmployeeExitInstructionRequest,
    _user: User = Depends(require_support_or_admin),
) -> EmployeeExitInstructionOut:
    text = build_employee_exit_instruction(payload.fio, payload.login, payload.password, payload.domain)
    return EmployeeExitInstructionOut(text=text)


@app.post("/api/ee_instruction/docx")
@app.post("/api/employee_exit/instruction/docx")
@app.post("/api/employee-exit/instruction/docx")
def employee_exit_instruction_docx(
    payload: EmployeeExitInstructionRequest,
    _user: User = Depends(require_support_or_admin),
) -> Response:
    content = build_employee_exit_instruction_docx_bytes(payload.fio, payload.login, payload.password, payload.domain)
    fname = attachment_filename_docx(payload.fio.strip() or "employee")
    ascii_fallback = re.sub(r"[^A-Za-z0-9._-]+", "_", fname) or "instrukciya.docx"
    cd = f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{quote(fname)}"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": cd},
    )


@app.get("/api/admin/users")
def admin_list_users(current_user: User = Depends(require_admin), db=Depends(get_db)) -> list[UserOut]:
    users = db.execute(select(User).where(User.role == "support").order_by(User.id)).scalars().all()
    return [
        UserOut(
            id=u.id,
            username=u.username,
            full_name=u.full_name,
            role=u.role,
            is_active_for_duties=bool(u.is_active_for_duties),
        )
        for u in users
    ]


@app.post("/api/admin/users", response_model=UserOut)
def admin_create_user(payload: CreateSupportUserRequest, current_user: User = Depends(require_admin), db=Depends(get_db)) -> UserOut:
    user = User(username=payload.username, full_name=payload.full_name, role="support", password_hash=hash_password(payload.password))
    db.add(user)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Username already exists")
    db.refresh(user)
    return UserOut(
        id=user.id,
        username=user.username,
        full_name=user.full_name,
        role=user.role,
        is_active_for_duties=bool(user.is_active_for_duties),
    )


@app.patch("/api/admin/users/{user_id}/duty-status", response_model=UserOut)
def admin_update_user_duty_status(
    user_id: int,
    payload: AdminDutyStatusRequest,
    current_user: User = Depends(require_admin),
    db=Depends(get_db),
) -> UserOut:
    user = ensure_support_user(db, user_id)
    user.is_active_for_duties = payload.is_active_for_duties
    db.add(user)
    db.commit()
    db.refresh(user)
    return UserOut(
        id=user.id,
        username=user.username,
        full_name=user.full_name,
        role=user.role,
        is_active_for_duties=bool(user.is_active_for_duties),
    )


@app.patch("/api/admin/users/{user_id}", response_model=UserOut)
def admin_update_user_profile(
    user_id: int,
    payload: AdminUpdateUserRequest,
    current_user: User = Depends(require_admin),
    db=Depends(get_db),
) -> UserOut:
    user = ensure_support_user(db, user_id)
    user.username = payload.username.strip()
    user.full_name = payload.full_name.strip()
    db.add(user)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Username already exists")
    db.refresh(user)
    return UserOut(
        id=user.id,
        username=user.username,
        full_name=user.full_name,
        role=user.role,
        is_active_for_duties=bool(user.is_active_for_duties),
    )


@app.post("/api/admin/users/{user_id}/password")
def admin_change_user_password(
    user_id: int,
    payload: AdminChangePasswordRequest,
    current_user: User = Depends(require_admin),
    db=Depends(get_db),
) -> dict:
    user = ensure_support_user(db, user_id)
    user.password_hash = hash_password(payload.new_password)
    db.add(user)
    db.commit()
    return {"ok": True, "user_id": user.id}


@app.delete("/api/admin/users/{user_id}/reports")
def admin_delete_user_reports(user_id: int, current_user: User = Depends(require_admin), db=Depends(get_db)) -> dict:
    user = ensure_support_user(db, user_id)
    reports = db.execute(
        select(DailyReport.id, DailyReport.date).where(DailyReport.support_user_id == user.id)
    ).all()
    report_ids = [int(report_id) for report_id, _ in reports]

    deleted_entries = 0
    deleted_reports = 0
    if report_ids:
        deleted_entries = int(
            db.execute(delete(ReportEntry).where(ReportEntry.report_id.in_(report_ids))).rowcount or 0
        )
        deleted_reports = int(
            db.execute(delete(DailyReport).where(DailyReport.id.in_(report_ids))).rowcount or 0
        )
        db.commit()

    deleted_exports = 0
    root = exports_dir().resolve()
    for report_id, report_date in reports:
        filename = f"report_{report_id}_{report_date.isoformat()}.xlsx"
        out_path = (root / filename).resolve()
        if out_path.is_file() and str(out_path).startswith(str(root)):
            try:
                out_path.unlink()
                deleted_exports += 1
            except OSError:
                pass

    return {
        "ok": True,
        "user_id": user.id,
        "deleted_reports": deleted_reports,
        "deleted_entries": deleted_entries,
        "deleted_exports": deleted_exports,
    }


@app.delete("/api/admin/users/{user_id}")
def admin_delete_user(user_id: int, current_user: User = Depends(require_admin), db=Depends(get_db)) -> dict:
    user = ensure_support_user(db, user_id)
    deleted_user_id = user.id
    db.delete(user)
    db.commit()
    return {"ok": True, "deleted_user_id": deleted_user_id}


@app.get("/api/duties", response_model=DutiesOut)
def get_duties(
    date_: date = Query(..., alias="date"),
    db=Depends(get_db),
    current_user: User = Depends(require_support_or_admin),
) -> DutiesOut:
    assignments = db.execute(
        select(DutyAssignment, User)
        .join(User, DutyAssignment.support_user_id == User.id)
        .where(DutyAssignment.date == date_)
    ).all()
    by_slot: dict[int, User] = {slot: user for (assignment, user) in assignments for slot in [assignment.slot]}

    slots_out: list[DutySlotOut] = []
    for slot in range(0, SLOT_COUNT):
        user = by_slot.get(slot)
        slots_out.append(
            DutySlotOut(
                slot=slot,
                start_time=slot_start_time_str(slot),
                user=UserOut(id=user.id, username=user.username, full_name=user.full_name, role=user.role) if user else None,
            )
        )
    return DutiesOut(date=date_, slots=slots_out)


@app.post("/api/duties/generate", response_model=DutiesGenerateOut)
def generate_duties(payload: DutiesGenerateRequest, current_user: User = Depends(require_admin), db=Depends(get_db)) -> DutiesGenerateOut:
    if payload.end_date < payload.start_date:
        raise HTTPException(status_code=400, detail="end_date must be >= start_date")

    support_users = db.execute(
        select(User).where(User.role == "support", User.is_active_for_duties == True).order_by(User.id)  # noqa: E712
    ).scalars().all()
    if not support_users:
        raise HTTPException(status_code=400, detail="No active support users configured")

    user_ids = [u.id for u in support_users]

    # Prefetch existing assignments in the range
    existing_in_range: dict[tuple[date, int], int] = {}
    existing_rows = db.execute(
        select(DutyAssignment).where(DutyAssignment.date >= payload.start_date, DutyAssignment.date <= payload.end_date)
    ).scalars().all()
    for row in existing_rows:
        existing_in_range[(row.date, row.slot)] = row.support_user_id

    # counts prior to start_date
    counts = {
        uid: 0
        for uid in user_ids
    }
    grouped = db.execute(
        select(DutyAssignment.support_user_id, func.count())
        .where(DutyAssignment.date < payload.start_date)
        .group_by(DutyAssignment.support_user_id)
    ).all()
    for uid, cnt in grouped:
        if uid in counts:
            counts[uid] = int(cnt)

    created = 0
    rng = random.SystemRandom()

    with db_session() as tx_db:
        if payload.overwrite:
            tx_db.execute(
                delete(DutyAssignment).where(
                    DutyAssignment.date >= payload.start_date,
                    DutyAssignment.date <= payload.end_date,
                )
            )
            tx_db.commit()

        # Recompute counts inside transaction to be consistent with any changes made above
        for uid in user_ids:
            counts[uid] = int(
                tx_db.execute(
                    select(func.count())
                    .select_from(DutyAssignment)
                    .where(DutyAssignment.support_user_id == uid, DutyAssignment.date < payload.start_date)
                ).scalar_one()
            )

        for day_index in range((payload.end_date - payload.start_date).days + 1):
            current_day = payload.start_date + timedelta(days=day_index)
            for slot in range(0, SLOT_COUNT):
                key = (current_day, slot)
                if (not payload.overwrite) and key in existing_in_range:
                    chosen_uid = existing_in_range[key]
                    counts[chosen_uid] = counts.get(chosen_uid, 0) + 1
                    continue

                min_count = min(counts.values())
                candidates = [uid for uid, c in counts.items() if c == min_count]
                chosen_uid = rng.choice(candidates)

                tx_db.add(DutyAssignment(date=current_day, slot=slot, support_user_id=chosen_uid))
                counts[chosen_uid] = counts.get(chosen_uid, 0) + 1
                created += 1

        tx_db.commit()

    return DutiesGenerateOut(
        start_date=payload.start_date,
        end_date=payload.end_date,
        overwrite=payload.overwrite,
        created_assignments=created,
    )


@app.post("/api/duties/batch")
def duties_batch(payload: DutiesBatchRequest, current_user: User = Depends(require_admin), db=Depends(get_db)) -> dict:
    if not payload.assignments:
        raise HTTPException(status_code=400, detail="assignments must not be empty")

    slots_seen: set[int] = set()
    for a in payload.assignments:
        if a.slot in slots_seen:
            raise HTTPException(status_code=400, detail="Duplicate slot in assignments")
        slots_seen.add(a.slot)

    # validate users
    user_ids = [a.user_id for a in payload.assignments]
    support_users = db.execute(select(User).where(User.id.in_(user_ids), User.role == "support")).scalars().all()
    support_user_ids = {u.id for u in support_users}
    missing = [uid for uid in user_ids if uid not in support_user_ids]
    if missing:
        raise HTTPException(status_code=400, detail=f"Some users are invalid: {missing[0]}")

    created = 0
    updated = 0
    with db_session() as tx_db:
        for a in payload.assignments:
            existing = tx_db.execute(
                select(DutyAssignment).where(DutyAssignment.date == payload.date, DutyAssignment.slot == a.slot)
            ).scalar_one_or_none()
            if existing:
                existing.support_user_id = a.user_id
                updated += 1
            else:
                tx_db.add(DutyAssignment(date=payload.date, slot=a.slot, support_user_id=a.user_id))
                created += 1
        tx_db.commit()

    return {"date": payload.date, "created": created, "updated": updated}


@app.post("/api/admin/notifications/duty-upcoming/dispatch", response_model=DutyNotificationDispatchOut)
def admin_dispatch_upcoming_duty_notification(
    at: Optional[datetime] = Query(None, description="Optional override time, ISO datetime"),
    current_user: User = Depends(require_admin),
    db=Depends(get_db),
) -> DutyNotificationDispatchOut:
    target_dt = (at or datetime.now()) + timedelta(minutes=2)
    duty_date, duty_slot = duty_slot_for_dt(target_dt)
    if duty_date is None or duty_slot is None:
        raise HTTPException(status_code=400, detail="No duty slot in +2 minute window")

    row = db.execute(
        select(DutyAssignment, User)
        .join(User, DutyAssignment.support_user_id == User.id)
        .where(DutyAssignment.date == duty_date, DutyAssignment.slot == duty_slot)
    ).first()
    if not row:
        return DutyNotificationDispatchOut(
            sent=False,
            reason="No assigned employee for upcoming slot",
            date=duty_date,
            slot=duty_slot,
            start_time=slot_start_time_str(duty_slot),
        )

    assignment, user = row
    payload = {
        "event": "duty_upcoming_2m",
        "date": assignment.date.isoformat(),
        "slot": int(assignment.slot),
        "start_time": slot_start_time_str(assignment.slot),
        "employee": {
            "id": int(user.id),
            "full_name": user.full_name,
            "username": user.username,
        },
    }

    if not settings.N8N_WEBHOOK_URL:
        return DutyNotificationDispatchOut(
            sent=False,
            reason="N8N webhook is not configured",
            date=assignment.date,
            slot=assignment.slot,
            start_time=slot_start_time_str(assignment.slot),
            employee_id=user.id,
            employee_name=user.full_name,
        )

    req = UrlRequest(
        settings.N8N_WEBHOOK_URL,
        data=json.dumps(payload).encode("utf-8"),
        method="POST",
        headers={"Content-Type": "application/json"},
    )
    try:
        with urlopen(req, timeout=settings.N8N_WEBHOOK_TIMEOUT_SEC):
            pass
    except HTTPError as e:
        raise HTTPException(status_code=502, detail=f"n8n webhook error: {e.code}")
    except URLError:
        raise HTTPException(status_code=502, detail="n8n webhook unavailable")

    return DutyNotificationDispatchOut(
        sent=True,
        date=assignment.date,
        slot=assignment.slot,
        start_time=slot_start_time_str(assignment.slot),
        employee_id=user.id,
        employee_name=user.full_name,
    )


@app.post("/api/duty-swaps", response_model=DutySwapOut)
def create_duty_swap_request(
    payload: DutySwapCreateRequest,
    current_user: User = Depends(get_current_user),
    db=Depends(get_db),
) -> DutySwapOut:
    if current_user.role != "support":
        raise HTTPException(status_code=403, detail="Only support users can create swap requests")
    if payload.from_slot == payload.to_slot:
        raise HTTPException(status_code=400, detail="Choose different duty slots")

    requester_assignment = db.execute(
        select(DutyAssignment).where(
            DutyAssignment.date == payload.date,
            DutyAssignment.slot == payload.from_slot,
        )
    ).scalar_one_or_none()
    if not requester_assignment or requester_assignment.support_user_id != current_user.id:
        raise HTTPException(status_code=400, detail="You are not assigned to the selected source slot")

    target_assignment = db.execute(
        select(DutyAssignment).where(
            DutyAssignment.date == payload.date,
            DutyAssignment.slot == payload.to_slot,
        )
    ).scalar_one_or_none()
    if not target_assignment:
        raise HTTPException(status_code=400, detail="Target slot has no assigned employee")

    target_user = ensure_support_user(db, int(target_assignment.support_user_id))
    if target_user.id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot create swap request with yourself")

    message = (
        f"{target_user.full_name}, {current_user.full_name} запрашивает обмен дежурствами "
        f"с {slot_start_time_str(payload.from_slot)} на {slot_start_time_str(payload.to_slot)}"
    )
    row = DutySwapRequest(
        date=payload.date,
        from_slot=payload.from_slot,
        to_slot=payload.to_slot,
        requester_user_id=current_user.id,
        target_user_id=target_user.id,
        message=message,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return DutySwapOut(
        id=row.id,
        date=row.date,
        from_slot=row.from_slot,
        to_slot=row.to_slot,
        requester_user_id=row.requester_user_id,
        target_user_id=row.target_user_id,
        message=row.message,
        status=row.status,
        created_at=row.created_at,
    )


@app.get("/api/duty-swaps/inbox", response_model=list[DutySwapOut])
def list_duty_swap_inbox(
    date_: Optional[date] = Query(None, alias="date"),
    current_user: User = Depends(get_current_user),
    db=Depends(get_db),
) -> list[DutySwapOut]:
    if current_user.role != "support":
        return []
    stmt = select(DutySwapRequest).where(DutySwapRequest.target_user_id == current_user.id)
    if date_ is not None:
        stmt = stmt.where(DutySwapRequest.date == date_)
    rows = db.execute(stmt.order_by(DutySwapRequest.created_at.desc())).scalars().all()
    return [
        DutySwapOut(
            id=r.id,
            date=r.date,
            from_slot=r.from_slot,
            to_slot=r.to_slot,
            requester_user_id=r.requester_user_id,
            target_user_id=r.target_user_id,
            message=r.message,
            status=r.status,
            created_at=r.created_at,
        )
        for r in rows
    ]


@app.post("/api/duty-swaps/{swap_id}/decision", response_model=DutySwapOut)
def decide_duty_swap_request(
    swap_id: int,
    payload: DutySwapDecisionRequest,
    current_user: User = Depends(get_current_user),
    db=Depends(get_db),
) -> DutySwapOut:
    if current_user.role != "support":
        raise HTTPException(status_code=403, detail="Only support users can decide swap requests")

    row = db.get(DutySwapRequest, swap_id)
    if not row:
        raise HTTPException(status_code=404, detail="Swap request not found")
    if row.target_user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Forbidden")
    if row.status != "pending":
        raise HTTPException(status_code=409, detail="Swap request is already processed")

    if payload.action == "accept":
        requester_slot = db.execute(
            select(DutyAssignment).where(
                DutyAssignment.date == row.date,
                DutyAssignment.slot == row.from_slot,
            )
        ).scalar_one_or_none()
        target_slot = db.execute(
            select(DutyAssignment).where(
                DutyAssignment.date == row.date,
                DutyAssignment.slot == row.to_slot,
            )
        ).scalar_one_or_none()
        if not requester_slot or not target_slot:
            raise HTTPException(status_code=409, detail="Duty slots are missing")
        if requester_slot.support_user_id != row.requester_user_id or target_slot.support_user_id != row.target_user_id:
            raise HTTPException(status_code=409, detail="Duty assignments changed, recreate swap request")

        requester_slot.support_user_id, target_slot.support_user_id = target_slot.support_user_id, requester_slot.support_user_id
        row.status = "accepted"
    else:
        row.status = "rejected"

    db.add(row)
    db.commit()
    db.refresh(row)
    return DutySwapOut(
        id=row.id,
        date=row.date,
        from_slot=row.from_slot,
        to_slot=row.to_slot,
        requester_user_id=row.requester_user_id,
        target_user_id=row.target_user_id,
        message=row.message,
        status=row.status,
        created_at=row.created_at,
    )


@app.post("/api/reports", response_model=DailyReportOut)
def create_or_get_report(payload: CreateOrGetReportRequest, current_user: User = Depends(get_current_user), db=Depends(get_db)) -> DailyReportOut:
    target_support_id: int
    if current_user.role == "support":
        target_support_id = current_user.id
        if payload.employee_id is not None and payload.employee_id != current_user.id:
            raise HTTPException(status_code=403, detail="Cannot create report for other employee")
    else:
        # admin
        if payload.employee_id is None:
            target_support_id = current_user.id
        else:
            target_support_id = payload.employee_id

    employee = ensure_support_user(db, target_support_id)

    existing = db.execute(
        select(DailyReport).where(DailyReport.date == payload.date, DailyReport.support_user_id == employee.id)
    ).scalar_one_or_none()

    if existing:
        # Ensure entries loaded for output
        db.refresh(existing)
        return report_to_out(db, existing)

    new_report = DailyReport(date=payload.date, support_user_id=employee.id, status="draft", finalized_at=None)
    db.add(new_report)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        existing = db.execute(
            select(DailyReport).where(DailyReport.date == payload.date, DailyReport.support_user_id == employee.id)
        ).scalar_one_or_none()
        if not existing:
            raise
        db.refresh(existing)
        return report_to_out(db, existing)
    db.refresh(new_report)
    return report_to_out(db, new_report)


@app.get("/api/reports", response_model=list[DailyReportOut])
def list_reports(
    date_: date = Query(..., alias="date"),
    employee_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db=Depends(get_db),
) -> list[DailyReportOut]:
    # role restrictions
    if current_user.role == "support":
        if employee_id is not None and employee_id != current_user.id:
            raise HTTPException(status_code=403, detail="Forbidden")
        employee_id = current_user.id
    else:
        # admin: if employee_id omitted -> all for that date
        if employee_id is not None:
            employee = ensure_support_user(db, employee_id)
            employee_id = employee.id

    stmt = select(DailyReport).where(DailyReport.date == date_)
    if employee_id is not None:
        stmt = stmt.where(DailyReport.support_user_id == employee_id)

    reports = db.execute(stmt.order_by(DailyReport.support_user_id)).scalars().all()
    out: list[DailyReportOut] = []
    for r in reports:
        db.refresh(r)
        out.append(report_to_out(db, r))
    return out


@app.put("/api/reports/{report_id}", response_model=DailyReportOut)
def update_report(
    report_id: int,
    payload: UpdateReportRequest,
    current_user: User = Depends(get_current_user),
    db=Depends(get_db),
) -> DailyReportOut:
    report = db.get(DailyReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    if current_user.role == "support" and report.support_user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Forbidden")

    new_date = payload.date or report.date
    new_support_id = payload.employee_id or report.support_user_id

    if current_user.role == "admin":
        if payload.employee_id is not None:
            ensure_support_user(db, payload.employee_id)
    else:
        # support can only update their own; ignore any provided fields
        new_date = report.date
        new_support_id = report.support_user_id

    # uniqueness check if admin changes date/support_user_id
    if (new_date != report.date) or (new_support_id != report.support_user_id):
        existing = db.execute(
            select(DailyReport).where(DailyReport.date == new_date, DailyReport.support_user_id == new_support_id, DailyReport.id != report.id)
        ).scalar_one_or_none()
        if existing:
            raise HTTPException(status_code=409, detail="Report for this employee/date already exists")

    # Update fields + replace entries
    report.date = new_date
    report.support_user_id = new_support_id

    db.execute(delete(ReportEntry).where(ReportEntry.report_id == report.id))
    for entry in payload.entries:
        db.add(ReportEntry(report_id=report.id, minutes=entry.minutes, description=entry.description))

    db.commit()
    # If report was finalized before edit, force Excel rebuild on next download/finalize.
    filename = report_excel_filename(report)
    out_path = (exports_dir() / filename).resolve()
    if out_path.exists():
        try:
            out_path.unlink()
        except OSError:
            pass
    db.refresh(report)
    return report_to_out(db, report)


def _filename_safe(filename: str) -> bool:
    if not filename or "/" in filename or "\\" in filename:
        return False
    return bool(re.fullmatch(r"[A-Za-z0-9._-]+", filename))


@app.post("/api/reports/{report_id}/finalize", response_model=ReportFinalizeOut)
def finalize_report(report_id: int, current_user: User = Depends(get_current_user), db=Depends(get_db)) -> ReportFinalizeOut:
    report = db.get(DailyReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    if current_user.role == "support" and report.support_user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Forbidden")

    # If already finalized, just return existing URL.
    if report.status == "final":
        filename = report_excel_filename(report)
        out_path = (exports_dir() / filename).resolve()
        if not out_path.exists():
            employee = db.get(User, report.support_user_id)
            if not employee:
                raise HTTPException(status_code=500, detail="Employee not found")
            build_report_excel(report, employee)
        return ReportFinalizeOut(report_id=report.id, status="final", excel_url=f"/exports/{filename}")

    employee = db.get(User, report.support_user_id)
    if not employee:
        raise HTTPException(status_code=500, detail="Employee not found")

    report.status = "final"
    report.finalized_at = datetime.utcnow()
    try:
        # Generate excel with final status and persist only after success.
        build_report_excel(report, employee)
        db.commit()
    except Exception:
        db.rollback()
        raise
    db.refresh(report)

    filename = report_excel_filename(report)
    return ReportFinalizeOut(report_id=report.id, status="final", excel_url=f"/exports/{filename}")


@app.get("/exports/{filename}")
def download_export(filename: str, current_user: User = Depends(get_current_user), db=Depends(get_db)) -> FileResponse:
    # Path traversal protection + strict naming rule.
    if not _filename_safe(filename):
        raise HTTPException(status_code=400, detail="Invalid filename")
    m = EXPORT_FILENAME_RE.match(filename)
    if not m:
        raise HTTPException(status_code=404, detail="Not found")

    report_id = int(m.group(1))
    report = db.get(DailyReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Not found")

    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")

    if report.status != "final":
        raise HTTPException(status_code=404, detail="Not finalized")

    out_path = (exports_dir() / filename).resolve()
    root = exports_dir().resolve()
    try:
        if not out_path.is_relative_to(root):  # py>=3.9
            raise HTTPException(status_code=400, detail="Invalid path")
    except AttributeError:  # pragma: no cover
        if str(root) not in str(out_path):
            raise HTTPException(status_code=400, detail="Invalid path")

    if not out_path.exists():
        employee = db.get(User, report.support_user_id)
        if not employee:
            raise HTTPException(status_code=500, detail="Employee not found")
        build_report_excel(report, employee)
        if not out_path.exists():
            raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(path=str(out_path), filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/admin/reports/export-all")
def admin_export_all_reports_excel(
    date_: date = Query(..., alias="date"),
    current_user: User = Depends(require_admin),
    db=Depends(get_db),
) -> StreamingResponse:
    support_users = db.execute(select(User).where(User.role == "support").order_by(User.id)).scalars().all()
    reports = db.execute(select(DailyReport).where(DailyReport.date == date_)).scalars().all()
    by_user_id = {r.support_user_id: r for r in reports}

    missing_employees: list[str] = []
    zip_buffer = io.BytesIO()
    used_names: set[str] = set()
    included = 0

    with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for user in support_users:
            report = by_user_id.get(user.id)
            if not report or report.status != "final":
                missing_employees.append(user.full_name)
                continue

            filename = report_excel_filename(report)
            out_path = (exports_dir() / filename).resolve()
            if not out_path.exists():
                build_report_excel(report, user)
            if not out_path.exists():
                missing_employees.append(user.full_name)
                continue

            surname = safe_export_name_part(surname_from_full_name(user.full_name))
            base_name = f"{surname} {date_.isoformat()}"
            arcname = f"{base_name}.xlsx"
            if arcname in used_names:
                arcname = f"{base_name} ({user.id}).xlsx"
            used_names.add(arcname)
            zf.writestr(arcname, out_path.read_bytes())
            included += 1

    if included == 0:
        raise HTTPException(status_code=400, detail="Нет сформированных Excel-файлов на выбранную дату")

    zip_buffer.seek(0)
    bundle_name = f"excel_reports_{date_.isoformat()}.zip"
    headers = {
        "Content-Disposition": f'attachment; filename="{bundle_name}"',
        "X-Missing-Employees": quote(",".join(missing_employees)),
        "X-Included-Count": str(included),
    }
    return StreamingResponse(zip_buffer, media_type="application/zip", headers=headers)
