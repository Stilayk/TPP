from __future__ import annotations

import io
import re
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from fastapi import HTTPException
from openpyxl import Workbook

from app.config import settings
from app.models import DailyReport, User
from app.schemas import DailyReportOut, ReportEntryOut, UserOut

try:
    from docx import Document
except ImportError:  # pragma: no cover
    Document = None  # type: ignore[misc, assignment]

EXPORT_FILENAME_RE = re.compile(r"^report_(\d+)_\d{4}-\d{2}-\d{2}\.xlsx$")


def is_safe_export_filename(filename: str) -> bool:
    if not filename or "/" in filename or "\\" in filename:
        return False
    return bool(re.fullmatch(r"[A-Za-z0-9._-]+", filename))


def exports_dir() -> Path:
    p = Path(settings.EXPORTS_DIR).expanduser()
    if not p.is_absolute():
        p = (Path.cwd() / p).resolve()
    p.mkdir(parents=True, exist_ok=True)
    return p


def build_employee_exit_instruction(fio: str, login: str, password: str, domain: str) -> str:
    fio = (fio or "").strip()
    login = (login or "").strip()
    password = (password or "").strip()
    domain = (domain or "").strip()
    domain_login_example = f"{domain}\\{login}"
    return (
        f"Добрый день, {fio}, я системный администратор в компании Sokolov, вам выдано оборудование.\n\n"
        "При включении ноутбука открывается bitlocker - стандартный пароль от него Sokolov2026 \n"
        f"Ваш логин - {login}, ваш пароль, при первом входе попросит сменить - {password}.\n\n"
        f"Ваш домен — {domain}.\n"
        f"Пример учётной записи в формате домена: {domain_login_example}\n\n"
        "Вход в сервисы осуществляется по доменной учётной записи.\n\n"
        "После входа в учетную запись вы можете войти в информационные ресурсы компании, почту, битрикс24. \n"
        "При входе в Битрикс у вас запросит адрес сайта - 'portal.hpdd.ru', логин и пароль от вашей доменной учётной записи.\n"
        "Для входа в Outlook также используется доменная учётная запись.\n"
        f"При входе в ZOOM нужно выбрать вход через Active Directory и ввести учётную запись в формате {domain_login_example} и пароль.\n\n"
        "Важно:\n"
        "• Папка 'Загрузки' автоматически очищается при перезагрузке.\n"
        "• Папка 'Документы' синхронизируется с сервером, для удобства при смене оборудования - данные будут синхронизированы.\n\n"
        "По всем вопросам вы можете набрать по номеру 8 800 1000 750 (добавочный уточняйте у руководителя или в службе поддержки)."
    )


def build_employee_exit_instruction_docx_bytes(fio: str, login: str, password: str, domain: str) -> bytes:
    if Document is None:
        raise HTTPException(
            status_code=503,
            detail="Генерация Word недоступна: в образе не установлен пакет python-docx. Пересоберите backend.",
        )
    text = build_employee_exit_instruction(fio, login, password, domain)
    doc = Document()
    for line in text.splitlines():
        doc.add_paragraph(line)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def safe_export_name_part(value: str) -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|]+", "_", value).strip()
    return cleaned or "Сотрудник"


def attachment_filename_docx(base_name: str) -> str:
    safe = safe_export_name_part(base_name) or "employee"
    return f"instrukciya_{safe}.docx"


def attachment_content_disposition_docx(filename: str) -> str:
    ascii_fallback = re.sub(r"[^A-Za-z0-9._-]+", "_", filename) or "instrukciya.docx"
    return f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{quote(filename)}"


def report_excel_filename(report: DailyReport) -> str:
    return f"report_{report.id}_{report.date.isoformat()}.xlsx"


def surname_from_full_name(full_name: str) -> str:
    parts = [p for p in (full_name or "").strip().split() if p]
    if not parts:
        return "Сотрудник"
    return parts[0]


def build_report_excel(report: DailyReport, user: User) -> Path:
    out_dir = exports_dir()
    filename = report_excel_filename(report)
    out_path = out_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"

    ws["A1"] = "Support Daily Report"
    ws["A3"] = "Employee"
    ws["B3"] = f"{user.full_name} ({user.username})"
    ws["A4"] = "Date"
    ws["B4"] = report.date.isoformat()
    ws["A5"] = "Status"
    ws["B5"] = report.status
    ws["A6"] = "GeneratedAt"
    ws["B6"] = datetime.utcnow().isoformat() + "Z"

    ws["A8"] = "№"
    ws["B8"] = "Minutes"
    ws["C8"] = "Description"

    total_minutes = 0
    row = 9
    for idx, entry in enumerate(report.entries, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = entry.minutes
        ws[f"C{row}"] = entry.description
        total_minutes += entry.minutes
        row += 1

    ws[f"A{row + 1}"] = "TotalMinutes"
    ws[f"B{row + 1}"] = total_minutes
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 80

    wb.save(out_path)
    return out_path


def report_to_out(db, report: DailyReport) -> DailyReportOut:
    employee = db.get(User, report.support_user_id)
    if not employee:
        raise HTTPException(status_code=500, detail="Employee not found")
    return DailyReportOut(
        report_id=report.id,
        date=report.date,
        employee_id=report.support_user_id,
        employee=UserOut(
            id=employee.id,
            username=employee.username,
            full_name=employee.full_name,
            role=employee.role,
        ),
        status=report.status,
        finalized_at=report.finalized_at,
        entries=[ReportEntryOut(minutes=e.minutes, description=e.description) for e in report.entries],
    )
