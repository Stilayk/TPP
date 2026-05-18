"""Выгрузка графика дежурств в Excel (период)."""

from __future__ import annotations

import io
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from app.duty_slots import SLOT_COUNT, slot_start_time_str
from app.models import DutyAssignment, User


def build_duties_period_excel_bytes(db, *, start_date: date, end_date: date) -> bytes:
    """Таблица: строки — календарные дни периода, столбцы — слоты; пустые ячейки «—»."""
    rows = db.execute(
        select(DutyAssignment.date, DutyAssignment.slot, User.full_name, User.username)
        .join(User, DutyAssignment.support_user_id == User.id)
        .where(DutyAssignment.date >= start_date, DutyAssignment.date <= end_date)
    ).all()

    by_cell: dict[tuple[date, int], str] = {}
    for d, slot, full_name, username in rows:
        label = ((full_name or "").strip() or (username or "").strip() or "—") if full_name or username else "—"
        by_cell[(d, int(slot))] = label

    wb = Workbook()
    ws = wb.active
    ws.title = "График"

    ws["A1"] = "Дата"
    for slot in range(SLOT_COUNT):
        col_letter = get_column_letter(2 + slot)
        ws[f"{col_letter}1"] = slot_start_time_str(slot)

    row_idx = 2
    d = start_date
    while d <= end_date:
        ws[f"A{row_idx}"] = d.isoformat()
        for slot in range(SLOT_COUNT):
            col_letter = get_column_letter(2 + slot)
            ws[f"{col_letter}{row_idx}"] = by_cell.get((d, slot), "—")
        row_idx += 1
        d += timedelta(days=1)

    ws.column_dimensions["A"].width = 12
    for slot in range(SLOT_COUNT):
        ws.column_dimensions[get_column_letter(2 + slot)].width = 22

    for r in ws.iter_rows(min_row=1, max_row=row_idx - 1, min_col=1, max_col=1 + SLOT_COUNT):
        for c in r:
            c.alignment = Alignment(vertical="center", wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
